"""
POI TC 자동 생성기
- Claude API로 TC 생성
- 그룹별 Excel 시트로 자동 분류 저장
"""

import anthropic
import json
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ================================
# 설정
# ================================
API_KEY = "여기에_API_키_입력"
ADAS_FUNCTION = "POI 검색 상세정보 수신 실패 시 동작 변경 사양 검증"
OUTPUT_FILE = "POI_TC.xlsx"

# ================================
# 프롬프트
# ================================
SYSTEM_PROMPT = """
너는 자동차 내비게이션 소프트웨어의 QC 엔지니어이며 테스트 설계 전문가야.
출력은 반드시 아래 JSON 형식만 사용해. 다른 텍스트는 절대 포함하지 마.

{
  "groups": [
    {
      "group_name": "그룹명",
      "testcases": [
        {
          "tc_id": "TC-001",
          "category": "정상/경계/부정/회귀/성능",
          "test_purpose": "테스트 목적",
          "precondition": "사전 조건",
          "procedure": "절차1\\n절차2\\n절차3",
          "input": "입력값/조건",
          "expected_result": "기대 결과",
          "note": "비고"
        }
      ]
    }
  ]
}
"""

USER_PROMPT = """
[역할 부여]
너는 자동차 내비게이션 소프트웨어의 QC 엔지니어이며, 테스트 설계 전문가야.

[상황 설명]
아래와 같이 내비게이션 POI 검색 상세정보 수신 실패 시의 동작 사양이 변경되었어.
이 변경 사양을 검증하기 위한 Testcase 시나리오를 설계해줘.

[사전 동작 조건]
- 사용자가 내비 POI 검색 수행
- 서버에 상세정보 요청
- 요청 후 5초 이내 응답이 없으면 실패로 간주

[변경 사양 정보]
- AS-IS(현재): "주소 정보 없음" 문구 표시, "목적지로/경유지로" 버튼 비활성화
- TO-BE(변경): 행정계 정보 또는 경위도 좌표 표시, "목적지로/경유지로" 버튼 비활성화.
  이후 서버 응답 수신 시 상세정보로 자동 업데이트 됨

변경 사양을 완전히 검증하기 위한 그룹을 네가 직접 판단해서 자유롭게 구성해줘.
그룹 수와 이름에 제한은 없으며, 아래 조건을 만족해야 해:
- 변경 사양의 모든 동작을 빠짐없이 커버할 것
- 엣지케이스, 경계값, 예외 상황 반드시 포함할 것
- 각 그룹당 최소 3개 이상 TC 작성
반드시 JSON 형식으로만 출력해.
"""

# ================================
# Excel 스타일
# ================================
GROUP_COLORS = ["4472C4", "ED7D31", "70AD47", "FFC000", "7030A0", "C00000", "00B0F0", "FF69B4"]

CATEGORY_COLORS = {
    "정상": "DDEEFF",
    "경계": "FFF2CC",
    "부정": "FFE0E0",
    "회귀": "E8F5E9",
    "성능": "F3E5F5",
}

HEADERS = ["TC-ID", "분류", "테스트 목적", "사전 조건", "테스트 절차", "입력값/조건", "기대 결과", "비고"]
COL_WIDTHS = [10, 8, 25, 25, 35, 20, 30, 15]
INVALID_CHARS = ['\\', '/', '*', '?', ':', '[', ']']


def clean_sheet_title(name):
    result = name
    for ch in INVALID_CHARS:
        result = result.replace(ch, '')
    return result[:30]


def get_thin_border():
    thin = Side(style='thin')
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def apply_header_style(cell, color_hex):
    cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
    cell.font = Font(bold=True, color="FFFFFF", size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = get_thin_border()


def apply_data_style(cell, bg_color=None):
    if bg_color:
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = get_thin_border()
    cell.font = Font(size=9)


def generate_tc_from_claude():
    print("Claude API 호출 중...")
    client = anthropic.Anthropic(api_key=API_KEY)
    message = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=20000,
        temperature=0.2,
        system=SYSTEM_PROMPT,
        messages=[{"role": "user", "content": USER_PROMPT}]
    )
    raw = message.content[0].text
    print(f"응답 수신 완료 (토큰: {message.usage.output_tokens})")
    return raw


def parse_tc_json(raw_text):
    match = re.search(r'\{.*\}', raw_text, re.DOTALL)
    if not match:
        raise ValueError("JSON 형식을 찾을 수 없습니다.")
    data = json.loads(match.group())
    print(f"파싱 완료: {len(data['groups'])}개 그룹")
    return data


def create_excel(data):
    wb = Workbook()
    ws_all = wb.active
    ws_all.title = "전체 TC"

    for col, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        cell = ws_all.cell(row=1, column=col, value=header)
        apply_header_style(cell, "2F4F4F")
        ws_all.column_dimensions[get_column_letter(col)].width = width
    ws_all.row_dimensions[1].height = 25

    all_row = 2
    total_tc = 0

    for g_idx, group in enumerate(data["groups"]):
        group_name = group["group_name"]
        testcases = group["testcases"]
        group_color = GROUP_COLORS[g_idx % len(GROUP_COLORS)]
        print(f"  그룹 생성 중: {group_name} ({len(testcases)}개)")

        safe_title = clean_sheet_title(group_name)
        ws = wb.create_sheet(title=safe_title)

        for col, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
            cell = ws.cell(row=1, column=col, value=header)
            apply_header_style(cell, group_color)
            ws.column_dimensions[get_column_letter(col)].width = width
        ws.row_dimensions[1].height = 25

        ws.insert_rows(1)
        ws.merge_cells(f"A1:{get_column_letter(len(HEADERS))}1")
        title_cell = ws["A1"]
        title_cell.value = f"【{group_name}】"
        title_cell.fill = PatternFill(start_color=group_color, end_color=group_color, fill_type="solid")
        title_cell.font = Font(bold=True, color="FFFFFF", size=12)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        sheet_row = 3

        for tc in testcases:
            category = tc.get("category", "")
            bg_color = None
            for key, color in CATEGORY_COLORS.items():
                if key in category:
                    bg_color = color
                    break

            row_data = [
                tc.get("tc_id", ""),
                tc.get("category", ""),
                tc.get("test_purpose", ""),
                tc.get("precondition", ""),
                tc.get("procedure", "").replace("\\n", "\n"),
                tc.get("input", ""),
                tc.get("expected_result", ""),
                tc.get("note", ""),
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=sheet_row, column=col, value=value)
                apply_data_style(cell, bg_color)
            ws.row_dimensions[sheet_row].height = 50
            sheet_row += 1

            for col, value in enumerate(row_data, 1):
                cell = ws_all.cell(row=all_row, column=col, value=value)
                apply_data_style(cell, bg_color)
            ws_all.row_dimensions[all_row].height = 50
            all_row += 1
            total_tc += 1

        ws.freeze_panes = "A3"

    ws_all.freeze_panes = "A2"

    ws_summary = wb.create_sheet(title="요약", index=1)
    ws_summary["A1"] = "TC 생성 요약"
    ws_summary["A1"].font = Font(bold=True, size=14)
    ws_summary["A3"] = "기능"
    ws_summary["B3"] = ADAS_FUNCTION
    ws_summary["A4"] = "총 TC 수"
    ws_summary["B4"] = total_tc
    ws_summary["A5"] = "그룹 수"
    ws_summary["B5"] = len(data["groups"])

    for i, group in enumerate(data["groups"]):
        ws_summary[f"A{6+i}"] = group["group_name"]
        ws_summary[f"B{6+i}"] = f"{len(group['testcases'])}개"

    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["B"].width = 30

    wb.save(OUTPUT_FILE)
    print(f"\nExcel 저장 완료: {OUTPUT_FILE}")
    print(f"총 {total_tc}개 TC, {len(data['groups'])}개 그룹")


if __name__ == "__main__":
    print("=== POI TC 자동 생성기 ===")
    print(f"대상 기능: {ADAS_FUNCTION}\n")

    try:
        raw = generate_tc_from_claude()
        data = parse_tc_json(raw)
        create_excel(data)
        print("\n완료! Excel 파일을 확인하세요.")

    except json.JSONDecodeError as e:
        print(f"JSON 파싱 실패: {e}")
        print("원본 응답:", raw[:500])
    except Exception as e:
        print(f"오류 발생: {e}")